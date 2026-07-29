import streamlit as st
from st_copy_to_clipboard import st_copy_to_clipboard
import re
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from datetime import datetime

st.set_page_config(layout="wide", page_title="AWP Format/Excel")

st.image("logo.png", width=150)

# Pull the main content up
st.markdown(
    """
    <style>
    .block-container { padding-top: 3rem; }
    </style>
    """,
    unsafe_allow_html=True
)

# Centered main heading
st.markdown(
    "<h1 style='text-align:center; margin-top:0;'>AWP Formatter</h1>",
    unsafe_allow_html=True
)


# =====================================================================
#  IMPORT TEMPLATE DEFINITION  (UU1AWP_10 Import Template.xlsx)
#  ---------------------------------------------------------------
#  The exported file must MATCH this template exactly:
#    Row 1  -> service header
#    Row 2  -> control row
#    Row 3  -> UDS field type    (one per column)
#    Row 4  -> internal field code (one per column)
#    Row 5  -> display label      (one per column)
#    Row 6  -> the data (a single AWP record)
#  Every one of the 149 template columns is always written, in the
#  template order. Columns the parser does not fill are left blank.
# =====================================================================

META_ROW1 = 'MP6441_ProcessUserDefinedScreenService_001#ADD'

META_ROW2 = [
    'AU1AWP',
    'CONTROL_ORG: ',
    'BUILD_ID: 12.3.1.1#PQJHPJSGMAQSDLSW_PRD',
    'WS_MAPPING:MP6441_ProcessUserDefinedScreenService_001=ADD;'
    'MP6441_ProcessUserDefinedScreenService_001=SYNC;'
    'MP6441_ProcessUserDefinedScreenService_001=DELETE;'
    'MP6441_ProcessUserDefinedScreenService_001=GET;',
]

# (field_type, internal_code, display_label) in exact template column order
TEMPLATE_COLUMNS = [
    ('UDSField#CHAR', 'AWP_CODE', 'AWP Number*'),
    ('UDSField#CHAR', 'AWP_STATUS', 'AWP Request Status*'),
    ('UDSField#CHAR', 'AWP_TYPE', 'AWP Type*'),
    ('UDSField#CHAR', 'DESCRIPTION', 'Description*'),
    ('UDSField#CHAR', 'AWP_ABCALCNO', 'ABC Ban Number and ALC Permit Number'),
    ('UDSField#CHKBOX', 'LOC_AOCC', 'AOCC/AOMF'),
    ('UDSField#CHKBOX', 'AWP_AIRFIELDACCESSACK', 'Accessing the Airfield Acknowledged\xa0'),
    ('UDSField#CHKBOX', 'WORK_AFTERNOON', 'Afternoon (1200-1859)'),
    ('UDSField#CHKBOX', 'LOC_BAY', 'Aircraft Bay'),
    ('UDSField#CHKBOX', 'ACC_AIRSIDEGATE', 'Airside Vehicle Gate'),
    ('UDSField#CHKBOX', 'LOC_ANCILLARY', 'Ancillary Building'),
    ('UDSField#CHAR', 'REQ_EMAIL', 'Applicants email'),
    ('UDSField#CHAR', 'REQ_PHONE', 'Applicants phone number'),
    ('UDSField#CHAR', 'AWP_APPROVALCOND', 'Approval Conditions or Requirements'),
    ('UDSField#CHKBOX', 'LOC_APRON', 'Apron'),
    ('UDSField#CHAR', 'AWP_ASSIGNEDTO', 'Assigned To for Review'),
    ('UDSField#CHAR', 'TOUR_WSIBOOKING', 'Authorised Tour booking WSI staff member'),
    ('UDSField#CHAR', 'TOUR_PICKUP', 'Bus Pick Up and Drop Off Location'),
    ('UDSField#CHKBOX', 'LOC_CARGO', 'Cargo Precinct'),
    ('UDSField#CHAR', 'REQ_SUPPLIERDESC', 'Company Description'),
    ('UDSField#CHAR', 'REQ_SUPPLIER', 'Company Name'),
    ('UDSField#CHAR', 'AWP_ASSET', 'Component ID / Location'),
    ('UDSField#CHKBOX', 'PER_CONFINED', 'Confined Space Sub Permit'),
    ('UDSField#CHKBOX', 'PER_CONTROLLEDACT', 'Controlled Activity Application - Cranes etc'),
    ('UDSField#CHKBOX', 'PER_CRANE', 'Crane Lift Sub Permit'),
    ('UDSField#DATE', 'AWP_CURRENTENDDATE', 'Current Exp date'),
    ('UDSField#CHKBOX', 'SD_DATA', 'Data'),
    ('UDSField#CHAR', 'SYS_OTHEDESC', 'Description (Other)'),
    ('UDSField#CHAR', 'AWP_LOCATIONDETAIL', 'Detailed Location of Works'),
    ('UDSField#CHAR', 'AWP_DETAILSCOPE', 'Detailed Scope of Works'),
    ('UDSField#CHAR', 'AWP_TAPSERVDETAIL', 'Details associated with tapping into the existing service.'),
    ('UDSField#CHAR', 'AWP_EQUIPDETAIL', 'Details for Required Equipment'),
    ('UDSField#CHAR', 'AWP_ABCALC', 'Do you have ABC and ALC approval?'),
    ('UDSField#CHAR', 'AWP_EQUIPREQU', 'Do you require WSI owned and managed equipment?'),
    ('UDSField#CHAR', 'AWP_ASSETINFO', 'Do you require any asset information?'),
    ('UDSField#CHKBOX', 'SD_ELEC', 'Electrical'),
    ('UDSField#CHKBOX', 'SYS_ELECHV', 'Electrical HV'),
    ('UDSField#CHKBOX', 'SYS_ELECLV', 'Electrical LV'),
    ('UDSField#CHKBOX', 'PER_EXCAVATION', 'Excavation & penetration Sub -Permit'),
    ('UDSField#CHAR', 'TOUR_EXTGUESSCAT', 'External Guest industry Category'),
    ('UDSField#CHKBOX', 'PER_FIRESYSTEM', 'Fire Isolation Permit (including smoke)'),
    ('UDSField#CHKBOX', 'SYS_FIRE', 'Fire Systems'),
    ('UDSField#CHKBOX', 'SD_FIRE', 'Fire detection system'),
    ('UDSField#CHKBOX', 'WORK_FRIDAY', 'Friday'),
    ('UDSField#CHKBOX', 'PER_GANTRY', 'Gantry Access Sub Permit'),
    ('UDSField#CHKBOX', 'LOC_GATELOUNGE', 'Gate Lounges'),
    ('UDSField#CHKBOX', 'SD_HVAC', 'HVAC'),
    ('UDSField#CHKBOX', 'SYS_HVAC', 'HVAC'),
    ('UDSField#CHAR', 'ACC_AVCH', 'Have you read the AVCH'),
    ('UDSField#CHKBOX', 'SD_HIGHVOLTAGE', 'High Voltage'),
    ('UDSField#CHKBOX', 'AWP_HOARDINGACK', 'Hoarding, Barricading\xa0 or Signage Acknowledged'),
    ('UDSField#CHAR', 'AWP_HOARDING', 'Hoarding, Barricading\xa0 or Signage Required'),
    ('UDSField#CHKBOX', 'PER_HOTWORK', 'Hot Work Sub Permit'),
    ('UDSField#CHAR', 'TOUR_ATTENDING', 'How many People will be on the Tour?'),
    ('UDSField#CHKBOX', 'SYS_HYDRAUL', 'Hydraulics'),
    ('UDSField#CHAR', 'AWP_OPSIMPACT', 'Impacts on Airport Ops'),
    ('UDSField#CHAR', 'AWP_OPSIMPACTS', 'Impacts on Airport Ops'),
    ('UDSField#CHKBOX', 'SDR_INSTALL', 'Installation'),
    ('UDSField#CHAR', 'TOUR_BUSREQU', 'Is a Bus Required?'),
    ('UDSField#CHKBOX', 'PER_ISOLATION', 'Isolation Sub Permit (combined, gas, electrical, stored, compressed air, water/s'),
    ('UDSField#CHKBOX', 'LOC_LANDSIDE', 'Landside'),
    ('UDSField#CHKBOX', 'AWP_LIGHTINGACK', 'Lighting Acknowledged'),
    ('UDSField#CHKBOX', 'ACC_LOADING', 'Loading Dock'),
    ('UDSField#CHKBOX', 'SDR_MAINT', 'Maintenance'),
    ('UDSField#CHAR', 'AWP_MATSTOREDDETAIL', 'Management plan for equipment, materials or chemical storage on site.'),
    ('UDSField#CHKBOX', 'PER_MATERIAL', 'Material Import Permit'),
    ('UDSField#CHAR', 'AWP_OPSMITIGATION', 'Mitigation Measures for Op impacts'),
    ('UDSField#CHKBOX', 'WORK_MONDAY', 'Monday'),
    ('UDSField#CHKBOX', 'WORK_MORNING', 'Morning (0500-1159)'),
    ('UDSField#CHKBOX', 'WORK_NIGHT', 'Night (1900-0459)'),
    ('UDSField#CHKBOX', 'PER_NA', 'Not Applicable'),
    ('UDSField#CHKBOX', 'SYS_NA', 'Not Applicable'),
    ('UDSField#CHKBOX', 'PER_OLS', 'OLS'),
    ('UDSField#CHKBOX', 'PER_OPSCLOSURE', 'Operational resource closure/shutdown sub-permit'),
    ('UDSField#CHAR', 'ACC_OTHER', 'Other Access Point'),
    ('UDSField#CHAR', 'LOC_OTHER', 'Other Location'),
    ('UDSField#CHAR', 'SDR_OTHER', 'Other Reason'),
    ('UDSField#CHAR', 'SD_OTHER', 'Other Shutdown/Isolation'),
    ('UDSField#CHKBOX', 'PER_CONSTRUCTOOH', 'Out of Hours Works'),
    ('UDSField#CHKBOX', 'PER_DISCHARGE', 'Permit to Discharge Water'),
    ('UDSField#CHKBOX', 'PER_ENTERAREA', 'Permit to Enter Protected Areas or No-Go Areas'),
    ('UDSField#CHAR', 'REQ_NAME', 'Person making the application'),
    ('UDSField#TIME', 'TOUR_COMMENCE', 'Planned Tour Commencement Time'),
    ('UDSField#DATE', 'AWP_ENDDATE', 'Proposed End Date'),
    ('UDSField#DATE', 'SD_ENDDATE', 'Proposed End date of the shutdown/Isolations'),
    ('UDSField#DATE', 'AWP_STARTDATE', 'Proposed Start Date'),
    ('UDSField#DATE', 'SD_STARTDATE', 'Proposed Start date of the shutdown/Isolations'),
    ('UDSField#TIME', 'TOUR_END', 'Proposed Tour End Time'),
    ('UDSField#DATE', 'AWP_PROPOSEDENDDATE', 'Proposed new Exp date'),
    ('UDSField#CHAR', 'AWP_TEMPSERVICEDETAIL', 'Provide Details of Temporary Services Required'),
    ('UDSField#CHAR', 'AWP_COMMSPLAN', 'Provide details of your communication plan.\xa0'),
    ('UDSField#CHKBOX', 'LOC_CARPARK', 'Public Carpark'),
    ('UDSField#CHAR', 'AWP_ABCALCREASON', 'Reason ABC and ALC Approval Not Required'),
    ('UDSField#CHAR', 'AWP_EXTENDREASON', 'Reason for Extension'),
    ('UDSField#CHKBOX', 'SDR_REPAIR', 'Repair'),
    ('UDSField#CHKBOX', 'PER_ROADOCCUPY', 'Road Occupancy'),
    ('UDSField#CHKBOX', 'AWP_ROADOCCUPYPLANACK', 'Road Occupancy and Traffice Management Plans Acknowledged'),
    ('UDSField#CHAR', 'AWP_ROADOCCUPYPLAN', 'Road Occupancy or Traffic Management Plans Required'),
    ('UDSField#CHKBOX', 'SYS_ROADSIGN', 'Roads and Signage'),
    ('UDSField#CHKBOX', 'PER_ROOFACCESS', 'Roof Access'),
    ('UDSField#CHKBOX', 'WORK_SATURDAY', 'Saturday'),
    ('UDSField#CHKBOX', 'SYS_SECURITY', 'Security'),
    ('UDSField#CHAR', 'AWP_EMERGNAME', 'Site Emergency/After Hours Contact Name'),
    ('UDSField#CHAR', 'AWP_EMERGPHONE', 'Site Emergency/After Hours Contact Number'),
    ('UDSField#CHKBOX', 'LOC_SITEWIDE', 'Site Wide'),
    ('UDSField#CHAR', 'ACC_SPCREQU', 'Special conditions or requirements associated with site Access'),
    ('UDSField#CHKBOX', 'AWP_SUBPERMITACK', 'Sub Permit Documentation Read and Understood'),
    ('UDSField#CHKBOX', 'WORK_SUNDAY', 'Sunday'),
    ('UDSField#CHKBOX', 'SYS_TECH', 'Technology & Network'),
    ('UDSField#CHKBOX', 'LOC_TERMARRIVAL', 'Terminal Arrivals'),
    ('UDSField#CHKBOX', 'LOC_TERMBAGROOM', 'Terminal Bag Room'),
    ('UDSField#CHKBOX', 'LOC_TERMBASEMENT', 'Terminal Basement'),
    ('UDSField#CHKBOX', 'LOC_TERMDEPART', 'Terminal Departures'),
    ('UDSField#CHKBOX', 'LOC_TERMLOADDOCK', 'Terminal Loading Dock'),
    ('UDSField#CHKBOX', 'ACC_TERMMAIN', 'Terminal Main Entry'),
    ('UDSField#CHKBOX', 'LOC_TERMROOF', 'Terminal Roof'),
    ('UDSField#CHKBOX', 'ACC_TERMSTAFF', 'Terminal Staff Entry'),
    ('UDSField#CHKBOX', 'SDR_TESTING', 'Testing'),
    ('UDSField#CHKBOX', 'WORK_THURSDAY', 'Thursday'),
    ('UDSField#CHAR', 'SD_DURATION', 'Total duration of the Shutdown/Isolations'),
    ('UDSField#CHKBOX', 'TOUR_ATTENDANCEFORM', 'Tour Attendance Form (Completed and Sent Document)'),
    ('UDSField#CHKBOX', 'WORK_TUESDAY', 'Tuesday'),
    ('UDSField#CHAR', 'AWP_WORKTYPE', 'Type of Work'),
    ('UDSField#CHAR', 'AWP_WORKTYPEOTHER', 'Type of Work (Other)'),
    ('UDSField#CHKBOX', 'PER_VEGETATION', 'Vegetation Works'),
    ('UDSField#CHKBOX', 'SYS_VTRANS', 'Vertical Transport'),
    ('UDSField#CHAR', 'TOUR_DEPARTMENT', 'WSI Division & Department Responsible for the Tour'),
    ('UDSField#CHAR', 'AWP_WSIREP', 'WSI Representative'),
    ('UDSField#CHAR', 'AWP_WASTEPLAN', 'Waste Management Plan Details'),
    ('UDSField#CHKBOX', 'SD_WATER', 'Water (potable/recycled/sewer etc)'),
    ('UDSField#CHKBOX', 'WORK_WEDNESDAY', 'Wednesday'),
    ('UDSField#CHAR', 'TOUR_BUSSIZE', 'What Size Bus is Required'),
    ('UDSField#CHAR', 'AWP_MATSTORED', 'Will Equipment, Materials or chemicals be stored onsite?'),
    ('UDSField#CHAR', 'AWP_TEMPSERVICE', 'Will Temporary Services be Required'),
    ('UDSField#CHAR', 'AWP_TAPSERV', 'Will the work include tapping into any existing services?\xa0'),
    ('UDSField#CHAR', 'SD_SHUTISOLATE', 'Will the work require any shutdowns and/or Isolations?'),
    ('UDSField#CHAR', 'AWP_LIGHTING', 'Will the work require lighting?'),
    ('UDSField#CHAR', 'AWP_TOOLS', 'Will tools be carried in and out of the Airport Terminal sterile areas?'),
    ('UDSField#CHAR', 'AWP_SUPERPHONE', 'Work hours site Supervisor Number'),
    ('UDSField#CHAR', 'AWP_SUPERNAME', 'Work hours site Supervisor name'),
    ('UDSField#CHKBOX', 'AWP_LANDSIDEACK', 'Working and Accessing Landside Areas Acknowledged\xa0'),
    ('UDSField#CHKBOX', 'PER_HEIGHT', 'Working at Height or Below Permit (including roof access permit)'),
    ('UDSField#CHKBOX', 'AWP_TERMINALWORKACK', 'Working in the Terminal Acknowledged\xa0'),
    ('UDSField#CHKBOX', 'AWP_AIRFIELDWORKACK', 'Working on the Airfield Acknowledged\xa0'),
    ('UDSField#DATI', 'AWP_CREATED', 'AWP Date Created'),
    ('UDSField#CHAR', 'AWP_ASSETDESC', 'Asset Description'),
    ('UDSField#CHAR', 'AWP_ASSETID', 'Asset ID'),
    ('UDSField#CHAR', 'AWP_WSIREPNAME', 'WSI Representative Name'),
    ('Comment#EN#$AWP#AWP_CODE', '', 'New Comment'),
]

# Maps each parser output (Section, Field) -> template internal code.
# Anything not listed here simply stays blank in the export.
FIELD_TO_CODE = {
    ("General", "Description"): "DESCRIPTION",
    ("General", "Approval Conditions"): "AWP_APPROVALCOND",
    ("General", "Company Name"): "REQ_SUPPLIER",
    ("Contact", "Person making the application"): "REQ_NAME",
    ("Contact", "Applicants Email"): "REQ_EMAIL",
    ("Contact", "Applicants Phone"): "REQ_PHONE",
    ("Contact", "WSI Rep"): "AWP_WSIREP",
    ("Approval", "ABC / ALC Approval"): "AWP_ABCALC",
    ("Approval", "ABC BAN Number / ALC Permit Number"): "AWP_ABCALCNO",
    ("Approval", "Reason Not Required"): "AWP_ABCALCREASON",
    ("Work", "Type of Work"): "AWP_WORKTYPE",
    ("Work", "Type of Work Other"): "AWP_WORKTYPEOTHER",
    ("Work", "Detailed Location of Works"): "AWP_LOCATIONDETAIL",
    ("Work", "Detailed Scope of Works"): "AWP_DETAILSCOPE",
    ("Work", "Start Date"): "AWP_STARTDATE",
    ("Work", "End Date"): "AWP_ENDDATE",
    ("Work", "Impacts on Airport Ops"): "AWP_OPSIMPACT",
    ("Work", "Mitigation Measures for Op impacts"): "AWP_OPSMITIGATION",
    ("Plans / Management", "Do you require and asset information?"): "AWP_ASSETINFO",
    ("Plans / Management", "Communication Plan"): "AWP_COMMSPLAN",
    ("Supervision", "Supervisor Name"): "AWP_SUPERNAME",
    ("Supervision", "Supervisor Phone"): "AWP_SUPERPHONE",
    ("Supervision", "Emergency Contact Name"): "AWP_EMERGNAME",
    ("Supervision", "Emergency Contact Phone"): "AWP_EMERGPHONE",
    ("Services", "Tools in Terminal sterile areas"): "AWP_TOOLS",
    ("Services", "Tapping Services"): "AWP_TAPSERV",
    ("Services", "Tapping Details"): "AWP_TAPSERVDETAIL",
    ("Acknowledgements", "Working on Airfield"): "AWP_AIRFIELDWORKACK",
    ("Acknowledgements", "Accessing Airfield"): "AWP_AIRFIELDACCESSACK",
    ("Acknowledgements", "Working Landside"): "AWP_LANDSIDEACK",
    ("Acknowledgements", "Working Terminal"): "AWP_TERMINALWORKACK",
    ("Waste Management", "Waste Management Plan"): "AWP_WASTEPLAN",
    ("Waste Management", "Equipment / Materials / Chemicals Stored On-Site"): "AWP_MATSTORED",
    ("Waste Management", "Management plan for equipment, materials or chemical storage on site"): "AWP_MATSTOREDDETAIL",
    ("Site Controls", "Hoarding / Barricading / Signage Required"): "AWP_HOARDING",
    ("Site Controls", "Hoarding / Barricading / Signage Acknowledged"): "AWP_HOARDINGACK",
    ("Site Controls", "Road Occupancy / Traffic Management Required"): "AWP_ROADOCCUPYPLAN",
    ("Temporary Services", "Temporary Services Required"): "AWP_TEMPSERVICE",
    ("Temporary Services", "Special Access Conditions (Personnel / Vehicles / Equipment)"): "ACC_SPCREQU",
    ("Location", "Terminal Departures"): "LOC_TERMDEPART",
    ("Location", "Terminal Arrivals"): "LOC_TERMARRIVAL",
    ("Location", "Terminal Basement"): "LOC_TERMBASEMENT",
    ("Location", "Terminal Loading Dock"): "LOC_TERMLOADDOCK",
    ("Location", "Terminal Bag Room"): "LOC_TERMBAGROOM",
    ("Location", "Gate Lounges"): "LOC_GATELOUNGE",
    ("Location", "Landside"): "LOC_LANDSIDE",
    ("Location", "Apron"): "LOC_APRON",
    ("Location", "Aircraft Bay"): "LOC_BAY",
    ("Location", "Cargo Precinct"): "LOC_CARGO",
    ("Location", "Public Carpark"): "LOC_CARPARK",
    ("Location", "AOCC/AOMF"): "LOC_AOCC",
    ("Location", "Terminal Roof"): "LOC_TERMROOF",
    ("Location", "Ancillary Building"): "LOC_ANCILLARY",
    ("Location", "Site-Wide"): "LOC_SITEWIDE",
    ("Location", "Other Area"): "LOC_OTHER",
    ("Access", "Airside Vehicle Gate"): "ACC_AIRSIDEGATE",
    ("Access", "Terminal Main Entry"): "ACC_TERMMAIN",
    ("Access", "Terminal Staff Entry"): "ACC_TERMSTAFF",
    ("Access", "Loading Dock"): "ACC_LOADING",
    ("Access", "Other Access"): "ACC_OTHER",
    ("Days", "Monday"): "WORK_MONDAY",
    ("Days", "Tuesday"): "WORK_TUESDAY",
    ("Days", "Wednesday"): "WORK_WEDNESDAY",
    ("Days", "Thursday"): "WORK_THURSDAY",
    ("Days", "Friday"): "WORK_FRIDAY",
    ("Days", "Saturday"): "WORK_SATURDAY",
    ("Days", "Sunday"): "WORK_SUNDAY",
    ("Hours", "Morning"): "WORK_MORNING",
    ("Hours", "Afternoon"): "WORK_AFTERNOON",
    ("Hours", "Night"): "WORK_NIGHT",
    ("Systems Impacted", "Not Applicable"): "SYS_NA",
    ("Systems Impacted", "Electrical LV"): "SYS_ELECLV",
    ("Systems Impacted", "Electrical HV"): "SYS_ELECHV",
    ("Systems Impacted", "HVAC"): "SYS_HVAC",
    ("Systems Impacted", "Technology & Network"): "SYS_TECH",
    ("Systems Impacted", "Security"): "SYS_SECURITY",
    ("Systems Impacted", "Hydraulics"): "SYS_HYDRAUL",
    ("Systems Impacted", "Roads and Signage"): "SYS_ROADSIGN",
    ("Systems Impacted", "Fire Systems"): "SYS_FIRE",
    ("Systems Impacted", "Vertical Transport"): "SYS_VTRANS",
    ("Systems Impacted", "Other System"): "SYS_OTHEDESC",
    ("Permits", "Not Applicable"): "PER_NA",
    ("Permits", "Confined Space Sub Permit"): "PER_CONFINED",
    ("Permits", "Out of Hours Works"): "PER_CONSTRUCTOOH",
    ("Permits", "Crane Lift Sub Permit"): "PER_CRANE",
    ("Permits", "Gantry Access Sub Permit"): "PER_GANTRY",
    ("Permits", "Excavation and Penetration Sub Permit"): "PER_EXCAVATION",
    ("Permits", "Hot Works Sub Permit"): "PER_HOTWORK",
    ("Permits", "Isolation Sub Permit"): "PER_ISOLATION",
    ("Permits", "Material Import Permit"): "PER_MATERIAL",
    ("Permits", "Operational Resource Closure/Shutdown Sub Permit"): "PER_OPSCLOSURE",
    ("Permits", "Road Occupancy"): "PER_ROADOCCUPY",
    ("Permits", "Permit to Discharge Water"): "PER_DISCHARGE",
    ("Permits", "Vegetation Works"): "PER_VEGETATION",
    ("Permits", "Working at Height or Below Permit"): "PER_HEIGHT",
    ("Permits", "Fire Isolation Permit"): "PER_FIRESYSTEM",
    ("Permits", "Permit to Enter Protected Areas or No-Go Areas"): "PER_ENTERAREA",
    ("Shutdown Required", "Shutdown Required"): "SD_SHUTISOLATE",
    ("Shutdown Types", "Electrical"): "SD_ELEC",
    ("Shutdown Types", "Data"): "SD_DATA",
    ("Shutdown Types", "HVAC"): "SD_HVAC",
    ("Shutdown Types", "Water"): "SD_WATER",
    ("Shutdown Types", "Fire detection system"): "SD_FIRE",
    ("Shutdown Types", "High Voltage"): "SD_HIGHVOLTAGE",
    ("Shutdown Types", "Other Shutdown Type"): "SD_OTHER",
    ("Reason For Shutdown", "Maintenance"): "SDR_MAINT",
    ("Reason For Shutdown", "Repair"): "SDR_REPAIR",
    ("Reason For Shutdown", "Installation"): "SDR_INSTALL",
    ("Reason For Shutdown", "Testing"): "SDR_TESTING",
    ("Reason For Shutdown", "Other Reason"): "SDR_OTHER",
    ("Shutdown Details", "Shutdown Start"): "SD_STARTDATE",
    ("Shutdown Details", "Shutdown End"): "SD_ENDDATE",
    ("Shutdown Details", "Shutdown Duration"): "SD_DURATION",
}


if st.session_state.get("clear_trigger"):
    st.session_state["raw_text"] = ""
    st.session_state["clear_trigger"] = False


type_col, number_col = st.columns([1, 1])

with number_col:
    awp_number = st.text_input(
        "Paste AWP Number",
        key="awp_number"
    )

with type_col:
    awp_type = st.selectbox(
        "AWP Type",
        [
            "New AWP",
            "Extend Existing AWP"
        ],
        key="awp_type"
    )

approval_conditions = st.text_area(
    "Approval Conditions or Requirements",
    height=120,
    placeholder="Type the approval conditions or requirements here...",
    key="approval_conditions"
)

raw = st.text_area(
    "Paste raw AWP data",
    height=250,
    key="raw_text"
)


col1, col2 = st.columns([1, 1])

with col1:
    process_clicked = st.button("Process", key="process_button", use_container_width=True)

with col2:
    if st.button("Clear", key="clear_button", use_container_width=True):
        st.session_state["clear_trigger"] = True
        st.session_state["data"] = None
        st.rerun()


# ---------- PROCESS ----------
if process_clicked:

    lines = [l.strip() for l in raw.split("\n") if l.strip()]
    data = {}
    key = None

    for line in lines:
        if ":" in line:
            k, v = line.split(":", 1)
            key = k.strip()
            data[key] = v.strip()
        elif key:
            data[key] += " " + line.strip()

    st.session_state["data"] = data


def to_output(raw_val, ftype):
    """
    Convert a parser value to the format the import template expects.

    CHKBOX columns  -> TRUE / FALSE / ""
    DATE columns    -> dd/mm/yyyy
    everything else -> plain text (standalone yes/no forced to YES/NO)
    """
    if raw_val is None:
        return ""

    s = str(raw_val).strip()

    # ---------------------------
    # CHECKBOX FIELDS
    # ---------------------------
    if ftype == "UDSField#CHKBOX":
        if "\u2611" in s:
            return "TRUE"
        if "\u2610" in s:
            return "FALSE"

        low = s.lower()

        if low in ("yes", "true"):
            return "TRUE"

        if low in ("no", "false"):
            return "FALSE"

        return ""

    # ---------------------------
    # DATE FIELDS
    # ---------------------------
    if ftype == "UDSField#DATE":
        if not s:
            return ""

        date_formats = [
            "%d/%m/%Y",
            "%d-%m-%Y",
            "%Y-%m-%d",
            "%Y/%m/%d",
            "%d/%m/%y",
            "%d-%m-%y"
        ]

        for fmt in date_formats:
            try:
                return datetime.strptime(s, fmt).strftime("%d/%m/%Y")
            except ValueError:
                pass

        return s

    # ---------------------------
    # CHAR / TIME / DATI
    # ---------------------------
    if "\u2611" in s or "\u2610" in s:
        if "\u2611 Yes" in s:
            return "YES"

        if "\u2611 No" in s:
            return "NO"

        return ""

    # Force any standalone yes/no answer to caps
    if s.lower() == "yes":
        return "YES"
    if s.lower() == "no":
        return "NO"

    return s


# ---------- EXCEL EXPORT FUNCTION ----------
def create_excel_file(export_rows):
    """Build an export that MATCHES the UU1AWP_10 import template.

    All 149 template columns are written in template order:
      row 1 -> service header
      row 2 -> control row
      row 3 -> UDS field type
      row 4 -> internal field code
      row 5 -> display label
      row 6 -> data (blank where the parser has no value)
    """
    # Collect parser values keyed by template code (last write wins).
    code_values = {}
    for row in export_rows:
        section = row.get("Section", "")
        field = row.get("Field", "")
        value = row.get("Value", "")
        if field == "" and value == "":
            continue  # skip section-heading rows
        code = FIELD_TO_CODE.get((section, field))
        if code:
            code_values[code] = value


    approval_cond = st.session_state.get("approval_conditions", "").strip()
    if approval_cond:
        code_values["AWP_APPROVALCOND"] = approval_cond


    # Populate AWP Number column/type...
    code_values["AWP_CODE"] = st.session_state.get("awp_number", "").strip()
    awp_type = st.session_state.get("awp_type", "")


    if awp_type == "New AWP":
        code_values["AWP_TYPE"] = "New AWP"

    elif awp_type == "Extend Existing AWP":
        code_values["AWP_TYPE"] = "Extend Existing AWP"

    wb = Workbook()
    ws = wb.active
    ws.title = "UU1AWP_10"

    # --- Row 1: service header ---
    ws.cell(row=1, column=1).value = META_ROW1

    # --- Row 2: control row ---
    for c, val in enumerate(META_ROW2, start=1):
        ws.cell(row=2, column=c).value = val

    # --- Rows 3-5 (type / code / label) and Row 6 (data) ---
    header_fill = PatternFill("solid", fgColor="1F77B4")
    header_font = Font(color="FFFFFF", bold=True)

    for c, (ftype, code, label) in enumerate(TEMPLATE_COLUMNS, start=1):
        ws.cell(row=3, column=c).value = ftype
        ws.cell(row=4, column=c).value = code

        label_cell = ws.cell(row=5, column=c)
        label_cell.value = label
        label_cell.fill = header_fill
        label_cell.font = header_font
        label_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Data row: blank if the parser produced nothing for this column.
        value_cell = ws.cell(row=6, column=c)
        if code and code in code_values:
            value_cell.value = to_output(code_values[code], ftype)
        else:
            value_cell.value = ""
        value_cell.alignment = Alignment(wrap_text=True, vertical="top")

        ws.column_dimensions[label_cell.column_letter].width = 28

    ws.freeze_panes = "A6"

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ---------- PARSER ----------
data = st.session_state.get("data")

if data:

    export_rows = []

    # Placeholder for the download button (rendered here, filled at the end)
    download_placeholder = st.empty()

    def add_section(section_name):
        export_rows.append({
            "Section": section_name,
            "Field": "",
            "Value": ""
        })

    def add_export(section, field, value):
        export_rows.append({
            "Section": section,
            "Field": field,
            "Value": value
        })

    def find_value(keyword):
        for k, v in data.items():
            if keyword.lower() in k.lower():
                return v
        return ""

    def find_other(keyword):
        for k, v in data.items():
            kl = k.lower()
            if "other" in kl and keyword.lower() in kl:
                return v
            if "if other" in kl and keyword.lower() in kl:
                return v
        return ""

    def find_shutdown_other():
        for k, v in data.items():
            kl = k.lower()
            if "other" in kl and ("shutdown" in kl or "isolation" in kl):
                return v
        return ""

    def check(val, src):
        val_clean = re.sub(r'[^a-z0-9\s]', '', val.lower())
        src_clean = re.sub(r'[^a-z0-9\s]', '', src.lower())

        return "\u2611" if val_clean in src_clean else "\u2610"

    def yesno(val):
        val = val.lower()
        if "yes" in val:
            return "\u2611 Yes    \u2610 No"
        if "no" in val:
            return "\u2610 Yes    \u2611 No"
        return "\u2610 Yes    \u2610 No"

    def ack(keyword):
        val = find_value(keyword)
        if not val:
            return "No"
        val = val.lower()
        if "true" in val or "yes" in val:
            return "Yes"
        if "false" in val or "no" in val:
            return "No"
        return val

    # ---------- SOURCES ----------
    area = find_value("located in")
    access = find_value("access")
    days = find_value("days")
    hours = find_value("working hours")
    systems = find_value("systems will be affected")
    temp_services = find_value("Temporary Services")
    access_conditions = find_value("special conditions")
    permits = find_value("What sub permits will be required throughout the duration of the works")
    shutdown_other = find_other("shutdown")
    shutdown = find_value("What type of Shutdown or Isolation is required")
    shutdown_main = find_value("require any shutdown")
    shutdown_reason = find_value("reason for the shutdown")
    combined_shutdown = shutdown + " " + shutdown_main
    shutdown_other = find_shutdown_other()
    other_area = find_other("area")
    other_access = find_other("access")
    combined_access = access + " " + other_access
    other_system = find_other("system")
    other_reason = find_other("reason")

    st.markdown("""
    <style>

    [data-theme="light"] .value {
        color: black;
        background: #f7f7f7;
    }

    [data-theme="dark"] .value {
        color: white;
        background: #333333;
    }

    .section {
        font-size:20px;
        font-weight:700;
        margin-top:30px;
        margin-bottom:10px;
        padding-bottom:6px;
        border-bottom:2px solid #e6e6e6;
        color:#1f77b4;
    }

    .label {
        font-weight:600;
        font-size:13px;
        margin-bottom:2px;
    }

    .value {
        font-family:monospace;
        font-size:13px;
        padding:6px 8px;
        border-radius:6px;
        margin-bottom:8px;
        border:1px solid #444;
    }

    </style>
    """, unsafe_allow_html=True)

    def field(label, keyword, i, section_name):
        val = find_value(keyword)
        add_export(section_name, label, val)

        col1, col2 = st.columns([6, 1])
        with col1:
            st.markdown(f"<div class='label'>{label}</div>", unsafe_allow_html=True)
            st.markdown(f"<div class='value'>{val}</div>", unsafe_allow_html=True)
        with col2:
            st_copy_to_clipboard(val, "Copy", "Done", key=f"{label}_{i}")

    # ---------- GENERAL ----------
    section = "General"
    add_section(section)
    st.markdown("<div class='section'>General</div>", unsafe_allow_html=True)

    for i, (l, k) in enumerate([
        ("Description", "Work/Activity Description"),
        ("Approval Conditions", "Approval Conditions"),
        ("Company Name", "Account"),
    ]):
        field(l, k, i, section)

    # ---------- CONTACT ----------
    section = "Contact"
    add_section(section)
    st.markdown("<div class='section'>Contact</div>", unsafe_allow_html=True)

    field("Person making the application", "Requested by", 100, section)
    field("Applicants Email", "Email address", 101, section)
    field("Applicants Phone", "Phone", 102, section)
    field("WSI Rep", "Who is your WSI representative", 103, section)

    # ---------- APPROVAL ----------
    section = "Approval"
    add_section(section)
    st.markdown("<div class='section'>Approval</div>", unsafe_allow_html=True)

    field("ABC / ALC Approval", "ABC", 50, section)
    field("ABC BAN Number / ALC Permit Number", "BAN", 51, section)
    field("Reason Not Required", "not applicable", 52, section)

    # ---------- WORK ----------
    section = "Work"
    add_section(section)
    st.markdown("<div class='section'>Work</div>", unsafe_allow_html=True)

    field("Type of Work", "Type of Work", 200, section)

    other_work = find_other("type of work")
    add_export(section, "Type of Work Other", other_work)

    col1, col2 = st.columns([6, 1])
    with col1:
        st.markdown("<div class='label'>Type of Work Other</div>", unsafe_allow_html=True)
        st.markdown(f"<div class='value'>{other_work}</div>", unsafe_allow_html=True)
    with col2:
        st_copy_to_clipboard(other_work, "Copy", "Done", key="type_of_work_other")

    field("Detailed Location of Works", "Detailed location", 202, section)
    field("Detailed Scope of Works", "Detailed scope", 203, section)
    field("Start Date", "start date", 204, section)
    field("End Date", "end date", 205, section)
    field("Impacts on Airport Ops", "Impacts", 206, section)
    field("Mitigation Measures for Op impacts", "Mitigation", 207, section)

    # ---------- PLANS ----------
    section = "Plans / Management"
    add_section(section)
    st.markdown("<div class='section'>Plans / Management</div>", unsafe_allow_html=True)

    field("Do you require and asset information?", "asset information", 300, section)
    field("Communication Plan", "communication plan", 301, section)

    # ---------- SUPERVISION ----------
    section = "Supervision"
    add_section(section)
    st.markdown("<div class='section'>Supervision</div>", unsafe_allow_html=True)

    field("Supervisor Name", "Supervisor name", 120, section)
    field("Supervisor Phone", "Supervisor phone", 121, section)
    field("Emergency Contact Name", "Site Emergency/After Hours Contact person name", 122, section)
    field("Emergency Contact Phone", "after hours contact person phone number", 123, section)

    # ---------- SERVICES ----------
    section = "Services"
    add_section(section)
    st.markdown("<div class='section'>Services</div>", unsafe_allow_html=True)

    field("Tools in Terminal sterile areas", "tools", 320, section)
    field("Tapping Services", "tapping", 321, section)
    field("Tapping Details", "details associated", 322, section)

    # ---------- ACKNOWLEDGEMENTS ----------
    section = "Acknowledgements"
    add_section(section)
    st.markdown("<div class='section'>Acknowledgements</div>", unsafe_allow_html=True)

    acknowledgements = [
        ("Working on Airfield", ack("airfield")),
        ("Accessing Airfield", ack("accessing the airfield")),
        ("Working Landside", ack("landside")),
        ("Working Terminal", ack("requirements around working in the Terminal")),
    ]

    for label, value in acknowledgements:
        st.markdown(f"{label}: {value}")
        add_export(section, label, value)

    # ---------- WASTE MANAGEMENT ----------
    section = "Waste Management"
    add_section(section)
    st.markdown("<div class='section'>Waste Management</div>", unsafe_allow_html=True)

    field("Waste Management Plan", "waste management", 302, section)

    storage_val = find_value("stored on-site")
    storage_result = yesno(storage_val)
    st.markdown(f"Equipment / Materials / Chemicals Stored On-Site: {storage_result}")
    add_export(section, "Equipment / Materials / Chemicals Stored On-Site", storage_result)

    field(
        "Management plan for equipment, materials or chemical storage on site",
        "management plan for equipment, materials or chemical storage",
        303,
        section
    )

    # ---------- SITE CONTROLS ----------
    section = "Site Controls"
    add_section(section)
    st.markdown("<div class='section'>Site Controls</div>", unsafe_allow_html=True)

    hoarding_val = find_value("Hoarding, Barricading")
    hoarding_result = yesno(hoarding_val)
    st.markdown(f"Hoarding / Barricading / Signage Required: {hoarding_result}")
    add_export(section, "Hoarding / Barricading / Signage Required", hoarding_result)

    hoarding_ack = ack("requirements around Hoarding, Barricading and Signage")
    st.markdown(f"Hoarding / Barricading / Signage Acknowledged: {hoarding_ack}")
    add_export(section, "Hoarding / Barricading / Signage Acknowledged", hoarding_ack)

    road_val = find_value("Road Occupancy or Traffic Management")
    road_result = yesno(road_val)
    st.markdown(f"Road Occupancy / Traffic Management Required: {road_result}")
    add_export(section, "Road Occupancy / Traffic Management Required", road_result)

    # ---------- TEMPORARY SERVICES ----------
    section = "Temporary Services"
    add_section(section)
    st.markdown("<div class='section'>Temporary Services</div>", unsafe_allow_html=True)

    temp_services_result = yesno(temp_services)
    st.markdown(f"Temporary Services Required: {temp_services_result}")
    add_export(section, "Temporary Services Required", temp_services_result)

    field(
        "Special Access Conditions (Personnel / Vehicles / Equipment)",
        "special conditions",
        340,
        section
    )

    # ---------- LOCATION ----------
    section = "Location"
    add_section(section)
    st.markdown("<div class='section'>Location</div>", unsafe_allow_html=True)

    locs = [
        "Terminal Departures", "Terminal Arrivals", "Terminal Basement",
        "Terminal Loading Dock", "Terminal Bag Room", "Gate Lounges",
        "Landside", "Apron", "Aircraft Bay", "Cargo Precinct",
        "Public Carpark", "AOCC/AOMF", "Terminal Roof",
        "Ancillary Building", "Site-Wide", "Other"
    ]

    c1, c2 = st.columns(2)
    for i, l in enumerate(locs):
        result = f"{check(l, area)} {l}"
        (c1 if i % 2 == 0 else c2).markdown(result)
        add_export(section, l, check(l, area))

    if check("Other", area) == "\u2611" and other_area:
        add_export(section, "Other Area", other_area)

        col1, col2 = st.columns([6, 1])
        with col1:
            st.markdown("<div class='label'>Other Area</div>", unsafe_allow_html=True)
            st.markdown(f"<div class='value'>{other_area}</div>", unsafe_allow_html=True)
        with col2:
            st_copy_to_clipboard(other_area, "Copy", "Done", key="other_area_copy")

    # ---------- ACCESS ----------
    section = "Access"
    add_section(section)
    st.markdown("<div class='section'>Access</div>", unsafe_allow_html=True)

    access_list = [
        "Airside Vehicle Gate", "Terminal Main Entry", "Terminal Staff Entry",
        "Loading Dock", "Other"
    ]

    c1, c2 = st.columns(2)
    for i, a in enumerate(access_list):
        result = f"{check(a, combined_access)} {a}"
        (c1 if i % 2 == 0 else c2).markdown(result)
        add_export(section, a, check(a, combined_access))

    if check("Other", access) == "\u2611" and other_access:
        add_export(section, "Other Access", other_access)

        col1, col2 = st.columns([6, 1])
        with col1:
            st.markdown("<div class='label'>Other Access</div>", unsafe_allow_html=True)
            st.markdown(f"<div class='value'>{other_access}</div>", unsafe_allow_html=True)
        with col2:
            st_copy_to_clipboard(other_access, "Copy", "Done", key="other_access_copy")

    # ---------- DAYS ----------
    section = "Days"
    add_section(section)
    st.markdown("<div class='section'>Days</div>", unsafe_allow_html=True)

    days_list = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]

    c1, c2 = st.columns(2)
    for i, d in enumerate(days_list):
        result = f"{check(d, days)} {d}"
        (c1 if i % 2 == 0 else c2).markdown(result)
        add_export(section, d, check(d, days))

    # ---------- HOURS ----------
    section = "Hours"
    add_section(section)
    st.markdown("<div class='section'>Hours</div>", unsafe_allow_html=True)

    for h in ["Morning", "Afternoon", "Night"]:
        result = f"{check(h, hours)} {h}"
        st.markdown(result)
        add_export(section, h, check(h, hours))

    # ---------- SYSTEMS ----------
    section = "Systems Impacted"
    add_section(section)
    st.markdown("<div class='section'>Systems Impacted</div>", unsafe_allow_html=True)

    systems_list = [
        "Electrical LV", "Electrical HV", "HVAC",
        "Technology & Network", "Security", "Hydraulics",
        "Roads and Signage", "Fire Systems", "Vertical Transport", "Other"
    ]

    not_applicable = "\u2611" if any(x in systems.lower() for x in ["na", "n/a", "not applicable"]) else "\u2610"
    st.markdown(f"{not_applicable} Not Applicable")
    add_export(section, "Not Applicable", not_applicable)

    for s in systems_list:
        result = f"{check(s, systems)} {s}"
        st.markdown(result)
        add_export(section, s, check(s, systems))

    if check("Other", systems) == "\u2611" and other_system:
        add_export(section, "Other System", other_system)

        col1, col2 = st.columns([6, 1])
        with col1:
            st.markdown("<div class='label'>Other System</div>", unsafe_allow_html=True)
            st.markdown(f"<div class='value'>{other_system}</div>", unsafe_allow_html=True)
        with col2:
            st_copy_to_clipboard(other_system, "Copy", "Done", key="other_system_copy")

    # ---------- PERMITS ----------
    section = "Permits"
    add_section(section)
    st.markdown("<div class='section'>Permits</div>", unsafe_allow_html=True)

    # (display label used for export mapping, distinctive keyword to match on)
    permits_list = [
        ("Confined Space Sub Permit", "confined"),
        ("Out of Hours Works", "out of hours"),
        ("Crane Lift Sub Permit", "crane"),
        ("Gantry Access Sub Permit", "gantry"),
        ("Excavation and Penetration Sub Permit", "excavation"),
        ("Hot Works Sub Permit", "hot work"),
        ("Isolation Sub Permit", "isolation sub"),
        ("Material Import Permit", "material import"),
        ("Operational Resource Closure/Shutdown Sub Permit", "operational resource"),
        ("Road Occupancy", "road occupancy"),
        ("Permit to Discharge Water", "discharge water"),
        ("Vegetation Works", "vegetation"),
        ("Working at Height or Below Permit", "height"),
        ("Fire Isolation Permit", "fire isolation"),
        ("Permit to Enter Protected Areas or No-Go Areas", "protected areas"),
    ]

    permits_na = "\u2611" if not permits.strip() else "\u2610"
    st.markdown(f"{permits_na} Not Applicable")
    add_export(section, "Not Applicable", permits_na)

    for label, keyword in permits_list:
        result = f"{check(keyword, permits)} {label}"
        st.markdown(result)
        add_export(section, label, check(keyword, permits))

    # ---------- SHUTDOWN ----------
    section = "Shutdown Required"
    add_section(section)
    st.markdown("<div class='section'>Shutdown Required</div>", unsafe_allow_html=True)

    shutdown_required_result = yesno(shutdown_main)
    st.markdown(shutdown_required_result)
    add_export(section, "Shutdown Required", shutdown_required_result)

    # ---------- SHUTDOWN TYPES ----------
    section = "Shutdown Types"
    add_section(section)
    st.markdown("<div class='section'>Shutdown Types</div>", unsafe_allow_html=True)

    for s in [
        "Electrical", "Data", "HVAC",
        "Water", "Fire detection system", "High Voltage", "Other"
    ]:
        result = f"{check(s, combined_shutdown)} {s}"
        st.markdown(result)
        add_export(section, s, check(s, combined_shutdown))

    if check("Other", shutdown) == "\u2611" and shutdown_other:
        add_export(section, "Other Shutdown Type", shutdown_other)

        col1, col2 = st.columns([6, 1])
        with col1:
            st.markdown("<div class='label'>Other Shutdown Type</div>", unsafe_allow_html=True)
            st.markdown(f"<div class='value'>{shutdown_other}</div>", unsafe_allow_html=True)
        with col2:
            st_copy_to_clipboard(shutdown_other, "Copy", "Done", key="other_shutdown_copy")

    # ---------- REASON FOR SHUTDOWN ----------
    section = "Reason For Shutdown"
    add_section(section)
    st.markdown("<div class='section'>Reason For Shutdown</div>", unsafe_allow_html=True)

    reason_list = [
        "Maintenance",
        "Repair",
        "Installation",
        "Testing",
        "Other"
    ]

    c1, c2 = st.columns(2)
    for i, r in enumerate(reason_list):
        result = f"{check(r, shutdown_reason)} {r}"
        (c1 if i % 2 == 0 else c2).markdown(result)
        add_export(section, r, check(r, shutdown_reason))

    if check("Other", shutdown_reason) == "\u2611" and other_reason:
        add_export(section, "Other Reason", other_reason)

        col1, col2 = st.columns([6, 1])
        with col1:
            st.markdown("<div class='label'>Other Reason</div>", unsafe_allow_html=True)
            st.markdown(f"<div class='value'>{other_reason}</div>", unsafe_allow_html=True)
        with col2:
            st_copy_to_clipboard(other_reason, "Copy", "Done", key="other_reason_reason_copy")

    # ---------- SHUTDOWN DETAILS ----------
    section = "Shutdown Details"
    add_section(section)
    st.markdown("<div class='section'>Shutdown Details</div>", unsafe_allow_html=True)

    field("Shutdown Start", "start date of the shutdown", 500, section)
    field("Shutdown End", "end date of the shutdown", 501, section)
    field("Shutdown Duration", "Total duration of the Shutdown/Isolations", 502, section)

    # ---------- DOWNLOAD EXCEL (rendered into the top placeholder) ----------
    excel_file = create_excel_file(export_rows)

    download_placeholder.download_button(
        label="Download Reformatted Data as Excel",
        data=excel_file.getvalue(),
        file_name="AWP_Reformatted_Data.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )
